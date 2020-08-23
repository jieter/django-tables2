import json
from datetime import date, datetime, time
from tempfile import NamedTemporaryFile
from unittest import skipIf

import pytz
import yaml
from django.core.exceptions import ImproperlyConfigured
from django.shortcuts import render
from django.test import TestCase
from openpyxl import load_workbook

import django_tables2 as tables
from django_tables2 import A
from django_tables2.config import RequestConfig

from .app.models import Occupation, Person, Region
from .utils import build_request

try:
    from django_tables2.export.export import TableExport
    from django_tables2.export.views import ExportMixin
except ImproperlyConfigured:
    TableExport = None


NAMES = [("Yildiz", "van der Kuil"), ("Lindi", "Hakvoort"), ("Gerardo", "Castelein")]
NAMES_LIST_OF_DICTS = [
    {"first_name": first_name, "last_name": last_name} for first_name, last_name in NAMES
]

CSV_SEP = "\r\n"

EXPECTED_CSV_DATA = tuple(",".join(name) for name in NAMES)
EXPECTED_CSV = CSV_SEP.join(("First name,Surname",) + EXPECTED_CSV_DATA) + CSV_SEP

EXPECTED_JSON = list(
    [{"First name": first_name, "Surname": last_name} for first_name, last_name in NAMES]
)


class Table(tables.Table):
    first_name = tables.Column()
    last_name = tables.Column()


class AccessorTable(tables.Table):
    given_name = tables.Column(accessor=A("first_name"), verbose_name="Given name")
    surname = tables.Column(accessor=A("last_name"))


class View(ExportMixin, tables.SingleTableView):
    table_class = Table
    table_pagination = {"per_page": 1}
    model = Person  # required for ListView
    template_name = "django_tables2/bootstrap.html"


@skipIf(TableExport is None, "Tablib is required to run the export tests")
class TableExportTest(TestCase):
    """
    github issue #474: null/None values in exports
    """

    def test_None_values(self):
        table = Table(
            [
                {"first_name": "Yildiz", "last_name": "van der Kuil"},
                {"first_name": "Jan", "last_name": None},
            ]
        )

        exporter = TableExport("csv", table)
        expected = ("First name,Last name", "Yildiz,van der Kuil", "Jan,")
        self.assertEqual(exporter.export(), CSV_SEP.join(expected) + CSV_SEP)

    def test_null_values(self):
        Person.objects.create(first_name="Jan", last_name="Coen")

        class Table(tables.Table):
            first_name = tables.Column()
            last_name = tables.Column(verbose_name="Last name")
            occupation = tables.Column(verbose_name="Occupation")

        table = Table(Person.objects.all())
        exporter = TableExport("csv", table)
        expected = ("First name,Last name,Occupation", "Jan,Coen,")
        self.assertEqual(exporter.export(), CSV_SEP.join(expected) + CSV_SEP)

    def test_export_accessors_list_of_dicts(self):
        table = AccessorTable(NAMES_LIST_OF_DICTS)

        exporter = TableExport("csv", table)
        expected = ("Given name,Surname",) + EXPECTED_CSV_DATA
        self.assertEqual(exporter.export(), CSV_SEP.join(expected) + CSV_SEP)

    def test_export_accessors_queryset(self):
        programmer = Occupation.objects.create(name="Programmer")

        for first_name, last_name in NAMES:
            Person.objects.create(first_name=first_name, last_name=last_name, occupation=programmer)

        class AccessorRelationTable(AccessorTable):
            occupation = tables.Column(accessor=A("occupation__name"), verbose_name="Occupation")

        table = AccessorRelationTable(Person.objects.all())

        exporter = TableExport("csv", table)
        expected = ("Given name,Surname,Occupation",) + tuple(
            row + "," + programmer.name for row in EXPECTED_CSV_DATA
        )
        self.assertEqual(exporter.export(), CSV_SEP.join(expected) + CSV_SEP)

    def test_export_dataset_kwargs(self):
        table = Table(
            [
                {"first_name": "Yildiz", "last_name": "van der Kuil"},
                {"first_name": "Jan", "last_name": None},
            ]
        )
        title = "My Custom Title"
        exporter = TableExport("xlsx", table, dataset_kwargs={"title": title})
        self.assertEqual(exporter.dataset.title, title)

    def test_export_default_dataset_title(self):
        class PersonTable(Table):
            class Meta:
                model = Person  # provides default title

        table = PersonTable(Person.objects.all())
        exporter = TableExport("xlsx", table)
        self.assertEqual(exporter.dataset.title, Person._meta.verbose_name_plural.title())


@skipIf(TableExport is None, "Tablib is required to run the export tests")
class ExportViewTest(TestCase):
    def setUp(self):
        for first_name, last_name in NAMES:
            Person.objects.create(first_name=first_name, last_name=last_name)

    def test_view_should_support_csv_export(self):
        response = View.as_view()(build_request("/?_export=csv"))
        self.assertEqual(response.getvalue().decode("utf8"), EXPECTED_CSV)

        # should just render the normal table without the _export query
        response = View.as_view()(build_request("/"))
        html = response.render().rendered_content

        self.assertIn("Yildiz", html)
        self.assertNotIn("Lindy", html)

    def test_should_raise_error_for_unsupported_file_type(self):
        table = Table([])

        with self.assertRaises(TypeError):
            TableExport(table=table, export_format="exe")

    def test_should_support_json_export(self):
        response = View.as_view()(build_request("/?_export=json"))
        self.assertEqual(json.loads(response.getvalue().decode("utf8")), EXPECTED_JSON)

    def test_should_support_yaml_export(self):
        response = View.as_view()(build_request("/?_export=yaml"))
        self.assertEqual(
            yaml.load(response.getvalue().decode("utf8"), Loader=yaml.FullLoader), EXPECTED_JSON
        )

    def test_should_support_custom_trigger_param(self):
        class View(ExportMixin, tables.SingleTableView):
            table_class = Table
            export_trigger_param = "export_to"
            model = Person  # required for ListView

        response = View.as_view()(build_request("/?export_to=json"))
        self.assertEqual(json.loads(response.getvalue().decode("utf8")), EXPECTED_JSON)

    def test_should_support_custom_filename(self):
        class View(ExportMixin, tables.SingleTableView):
            table_class = Table
            export_name = "people"
            model = Person  # required for ListView

        response = View.as_view()(build_request("/?_export=json"))
        self.assertEqual(response["Content-Disposition"], 'attachment; filename="people.json"')

    def test_function_view(self):
        """Test the code used in the docs."""

        def table_view(request):
            table = Table(Person.objects.all())
            RequestConfig(request).configure(table)

            export_format = request.GET.get("_export", None)
            if TableExport.is_valid_format(export_format):
                exporter = TableExport(export_format, table)
                return exporter.response("table.{}".format(export_format))

            return render(request, "django_tables2/table.html", {"table": table})

        response = table_view(build_request("/?_export=csv"))
        self.assertEqual(response.getvalue().decode("utf8"), EXPECTED_CSV)

        # must also support the normal html table.
        response = table_view(build_request("/"))
        html = response.content.decode("utf8")

        self.assertIn("Yildiz", html)
        self.assertNotIn("Lindy", html)

    def test_should_support_custom_dataset_kwargs(self):
        title = "The Sheet Name"

        class View(ExportMixin, tables.SingleTableView):
            table_class = Table
            model = Person  # required for ListView
            dataset_kwargs = {"title": title}

        response = View.as_view()(build_request("/?_export=xlsx"))
        self.assertEqual(response.status_code, 200)

        with NamedTemporaryFile(suffix=".xlsx") as tmp:
            tmp.write(response.content)
            tmp.seek(0)
            wb = load_workbook(tmp.name)
            self.assertIn(title, wb.sheetnames)


class OccupationTable(tables.Table):
    name = tables.Column()
    boolean = tables.Column()
    region = tables.Column()


class OccupationView(ExportMixin, tables.SingleTableView):
    model = Occupation
    table_class = OccupationTable
    table_pagination = {"per_page": 1}
    template_name = "django_tables2/bootstrap.html"


@skipIf(TableExport is None, "Tablib is required to run the export tests")
class AdvancedExportViewTest(TestCase):
    @classmethod
    def setUpClass(cls):
        super().setUpClass()

        richard = Person.objects.create(first_name="Richard", last_name="Queener")

        vlaanderen = Region.objects.create(name="Vlaanderen", mayor=richard)
        Occupation.objects.create(name="Timmerman", boolean=True, region=vlaanderen)
        Occupation.objects.create(name="Ecoloog", boolean=False, region=vlaanderen)

    def test_should_work_with_foreign_keys(self):
        response = OccupationView.as_view()(build_request("/?_export=xls"))
        data = response.content
        # binary data, so not possible to compare to an exact expectation
        self.assertTrue(data.find("Vlaanderen".encode()))
        self.assertTrue(data.find("Ecoloog".encode()))
        self.assertTrue(data.find("Timmerman".encode()))

    def test_datetime_xls(self):
        """Verify datatime objects can be exported to xls."""

        utc = pytz.timezone("UTC")

        class Table(tables.Table):
            date = tables.DateColumn()
            time = tables.TimeColumn()
            datetime = tables.DateTimeColumn()

        class View(ExportMixin, tables.SingleTableView):
            table_class = Table
            table_pagination = {"per_page": 1}
            template_name = "django_tables2/bootstrap.html"

            def get_queryset(self):
                return [
                    {
                        "date": date(2019, 7, 22),
                        "time": time(11, 11, 11),
                        "datetime": utc.localize(datetime(2019, 7, 22, 11, 11, 11)),
                    }
                ]

        response = View.as_view()(build_request("/?_export=csv"))
        data = response.getvalue().decode("utf8")

        expected_csv = "\r\n".join(
            ("Date,Time,Datetime", "2019-07-22,11:11:11,2019-07-22 13:11:11", "")
        )
        self.assertEqual(data, expected_csv)

        response = View.as_view()(build_request("/?_export=xls"))
        self.assertIn("2019-07-22 13:11:11".encode(), response.content)

    def test_export_invisible_columns(self):
        """Verify columns with visible=False *do* get exported."""

        DATA = [{"name": "Bess W. Fletcher", "website": "teammonka.com"}]

        class Table(tables.Table):
            name = tables.Column()
            website = tables.Column(visible=False)

        class View(ExportMixin, tables.SingleTableView):
            table_class = Table
            table_pagination = {"per_page": 1}
            template_name = "django_tables2/bootstrap.html"

            def get_queryset(self):
                return DATA

        response = View.as_view()(build_request())
        self.assertNotContains(response, "teammonka.com")

        response = View.as_view()(build_request("/?_export=csv"))

        data = response.getvalue().decode()

        expected_csv = "\r\n".join(("Name,Website", "Bess W. Fletcher,teammonka.com", ""))
        self.assertEqual(data, expected_csv)

    def test_should_work_with_foreign_key_fields(self):
        class OccupationWithForeignKeyFieldsTable(tables.Table):
            name = tables.Column()
            boolean = tables.Column()
            region = tables.Column()
            mayor = tables.Column(accessor="region__mayor__first_name")

        class View(ExportMixin, tables.SingleTableView):
            table_class = OccupationWithForeignKeyFieldsTable
            table_pagination = {"per_page": 1}
            model = Occupation
            template_name = "django_tables2/bootstrap.html"

        response = View.as_view()(build_request("/?_export=csv"))
        data = response.getvalue().decode("utf8")

        expected_csv = "\r\n".join(
            (
                "Name,Boolean,Region,First name",
                "Timmerman,True,Vlaanderen,Richard",
                "Ecoloog,False,Vlaanderen,Richard",
                "",
            )
        )
        self.assertEqual(data, expected_csv)

    def test_should_allow_exclude_columns(self):
        class OccupationExcludingView(ExportMixin, tables.SingleTableView):
            table_class = OccupationTable
            table_pagination = {"per_page": 1}
            model = Occupation
            template_name = "django_tables2/bootstrap.html"
            exclude_columns = ("boolean",)

        response = OccupationExcludingView.as_view()(build_request("/?_export=csv"))
        data = response.getvalue().decode("utf8")

        self.assertEqual(data.splitlines()[0], "Name,Region")


@skipIf(TableExport is None, "Tablib is required to run the export tests")
class UnicodeExportViewTest(TestCase):
    def test_exporting_unicode_data(self):
        unicode_name = "木匠"
        Occupation.objects.create(name=unicode_name)

        expected_csv = "Name,Boolean,Region\r\n{},,\r\n".format(unicode_name)

        response = OccupationView.as_view()(build_request("/?_export=csv"))
        self.assertEqual(response.getvalue().decode("utf8"), expected_csv)

        # smoke tests, hard to test this binary format for string containment
        response = OccupationView.as_view()(build_request("/?_export=xls"))
        self.assertGreater(len(response.content), len(expected_csv))

        response = OccupationView.as_view()(build_request("/?_export=xlsx"))
        self.assertGreater(len(response.content), len(expected_csv))

    def test_exporting_unicode_header(self):
        unicode_header = "hé"

        class Table(tables.Table):
            name = tables.Column(verbose_name=unicode_header)

        exporter = TableExport("csv", Table([]))
        response = exporter.response()
        self.assertEqual(response.getvalue().decode("utf8"), unicode_header + "\r\n")

        exporter = TableExport("xls", Table([]))
        # this would fail if the header contains unicode and string converstion is attempted.
        exporter.export()
